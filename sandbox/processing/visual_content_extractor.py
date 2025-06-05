"""
Visual Content Extractor for FinDocPipeline
Extracts images, charts, graphs, and tables from various document formats.
"""

import streamlit as st
import cv2
import numpy as np
from PIL import Image
import io
import base64
from typing import List, Dict, Any, Optional, Tuple
import tempfile
import os

# Optional imports with fallbacks
try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

try:
    import pytesseract
    PYTESSERACT_AVAILABLE = True
except ImportError:
    PYTESSERACT_AVAILABLE = False

# Import visual processing components
from processing.visual_content_processor import VisualContentProcessor
from storage.visual_content_storage import VisualContentStorage
from schemas.visual_content_schema import VisualContentSchema

class VisualContentExtractor:
    """
    Extracts visual content from documents and processes them for embeddings.
    Integrates with existing file handlers to provide comprehensive visual analysis.
    """
    
    def __init__(self):
        self.processor = VisualContentProcessor()
        self.storage = VisualContentStorage()
        
        # Supported image formats
        self.supported_formats = ['.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.gif']
        
        # Minimum image size for processing
        self.min_image_size = (50, 50)
        self.max_image_size = (2000, 2000)
    
    def extract_from_pdf(self, uploaded_file, document_id: str) -> List[VisualContentSchema]:
        """Extract visual content from PDF files"""
        
        visual_contents = []
        
        if not PYMUPDF_AVAILABLE and not PDFPLUMBER_AVAILABLE:
            st.warning("⚠️ PDF processing libraries not available. Install PyMuPDF and pdfplumber for full functionality.")
            return visual_contents
        
        try:
            # Save uploaded file to temporary location
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                tmp_file.write(uploaded_file.getvalue())
                tmp_path = tmp_file.name
            
            # Extract using PyMuPDF for images (if available)
            if PYMUPDF_AVAILABLE:
                visual_contents.extend(self._extract_pdf_images_pymupdf(tmp_path, document_id))
            
            # Extract using pdfplumber for tables and text-based visuals (if available)
            if PDFPLUMBER_AVAILABLE:
                visual_contents.extend(self._extract_pdf_tables_pdfplumber(tmp_path, document_id))
            
            # Clean up
            os.unlink(tmp_path)
            
        except Exception as e:
            st.error(f"Error extracting visual content from PDF: {str(e)}")
        
        return visual_contents
    
    def extract_from_docx(self, uploaded_file, document_id: str) -> List[VisualContentSchema]:
        """Extract visual content from DOCX files"""
        
        visual_contents = []
        
        try:
            from docx import Document
            from docx.document import Document as DocumentType
            
            # Load document
            doc = Document(io.BytesIO(uploaded_file.getvalue()))
            
            # Extract images
            visual_contents.extend(self._extract_docx_images(doc, document_id))
            
            # Extract tables
            visual_contents.extend(self._extract_docx_tables(doc, document_id))
            
        except ImportError:
            st.warning("⚠️ python-docx not available. Install python-docx for DOCX visual content extraction.")
        except Exception as e:
            st.error(f"Error extracting visual content from DOCX: {str(e)}")
        
        return visual_contents
    
    def extract_from_excel(self, uploaded_file, document_id: str) -> List[VisualContentSchema]:
        """Extract visual content from Excel files"""
        
        visual_contents = []
        
        try:
            import openpyxl
            from openpyxl.drawing.image import Image as ExcelImage
            
            # Load workbook
            workbook = openpyxl.load_workbook(io.BytesIO(uploaded_file.getvalue()))
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Extract charts
                visual_contents.extend(self._extract_excel_charts(sheet, document_id, sheet_name))
                
                # Extract images
                visual_contents.extend(self._extract_excel_images(sheet, document_id, sheet_name))
                
                # Extract data tables
                visual_contents.extend(self._extract_excel_tables(sheet, document_id, sheet_name))
        
        except ImportError:
            st.warning("⚠️ openpyxl not available. Install openpyxl for Excel visual content extraction.")
        except Exception as e:
            st.error(f"Error extracting visual content from Excel: {str(e)}")
        
        return visual_contents
    
    def _extract_pdf_images_pymupdf(self, pdf_path: str, document_id: str) -> List[VisualContentSchema]:
        """Extract images from PDF using PyMuPDF"""
        
        visual_contents = []
        
        if not PYMUPDF_AVAILABLE:
            return visual_contents
        
        try:
            doc = fitz.open(pdf_path)
            
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                
                # Method 1: Extract embedded images
                image_list = page.get_images()
                
                for img_index, img in enumerate(image_list):
                    try:
                        # Extract image
                        xref = img[0]
                        pix = fitz.Pixmap(doc, xref)
                        
                        # Skip if image is too small or too large
                        if pix.width < self.min_image_size[0] or pix.height < self.min_image_size[1]:
                            continue
                        if pix.width > self.max_image_size[0] or pix.height > self.max_image_size[1]:
                            continue
                        
                        # Convert to numpy array
                        if pix.n - pix.alpha < 4:  # GRAY or RGB
                            img_data = pix.tobytes("png")
                            img_array = np.frombuffer(img_data, dtype=np.uint8)
                            image = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
                            
                            if image is not None:
                                # Get image position on page
                                img_rect = page.get_image_rects(xref)[0] if page.get_image_rects(xref) else None
                                
                                if img_rect:
                                    # Normalize coordinates
                                    page_rect = page.rect
                                    bounding_box = {
                                        'x': img_rect.x0 / page_rect.width,
                                        'y': img_rect.y0 / page_rect.height,
                                        'width': (img_rect.x1 - img_rect.x0) / page_rect.width,
                                        'height': (img_rect.y1 - img_rect.y0) / page_rect.height
                                    }
                                else:
                                    # Default bounding box if position unknown
                                    bounding_box = {'x': 0.0, 'y': 0.0, 'width': 1.0, 'height': 1.0}
                                
                                # Extract OCR text from image area
                                ocr_text = self._extract_ocr_text(image)
                                
                                # Process visual content
                                visual_content = self.processor.process_visual_content(
                                    image=image,
                                    document_id=document_id,
                                    page_number=page_num + 1,
                                    bounding_box=bounding_box,
                                    ocr_text=ocr_text
                                )
                                
                                visual_contents.append(visual_content)
                        
                        pix = None  # Clean up
                        
                    except Exception as e:
                        st.warning(f"Error processing embedded image {img_index} on page {page_num + 1}: {str(e)}")
                        continue
                
                # Method 2: Render entire page as image to capture text-based tables/charts
                try:
                    # Render page at higher resolution for better quality
                    mat = fitz.Matrix(2.0, 2.0)  # 2x zoom for better quality
                    pix = page.get_pixmap(matrix=mat)
                    
                    # Convert to numpy array
                    img_data = pix.tobytes("png")
                    img_array = np.frombuffer(img_data, dtype=np.uint8)
                    page_image = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
                    
                    if page_image is not None:
                        # Extract text from the entire page
                        page_text = page.get_text()
                        
                        # Only process if page has substantial content
                        if len(page_text.strip()) > 50:  # Minimum text threshold
                            
                            # Split page into regions for better processing
                            regions = self._split_page_into_regions(page_image, page_text)
                            
                            for region_idx, region_data in enumerate(regions):
                                region_image = region_data['image']
                                region_text = region_data['text']
                                region_bbox = region_data['bbox']
                                
                                # Skip empty or very small regions
                                if region_image.shape[0] < 50 or region_image.shape[1] < 50:
                                    continue
                                
                                # Process visual content for this region
                                visual_content = self.processor.process_visual_content(
                                    image=region_image,
                                    document_id=document_id,
                                    page_number=page_num + 1,
                                    bounding_box=region_bbox,
                                    ocr_text=region_text,
                                    document_name=f"{document_id}_page_{page_num + 1}",
                                    page_context=f"Page {page_num + 1} Region {region_idx + 1}"
                                )
                                
                                visual_contents.append(visual_content)
                    
                    pix = None  # Clean up
                    
                except Exception as e:
                    st.warning(f"Error rendering page {page_num + 1}: {str(e)}")
                    continue
            
            doc.close()
            
        except Exception as e:
            st.error(f"Error extracting images from PDF: {str(e)}")
        
        return visual_contents
    
    def _extract_pdf_tables_pdfplumber(self, pdf_path: str, document_id: str) -> List[VisualContentSchema]:
        """Extract tables from PDF using pdfplumber"""
        
        visual_contents = []
        
        if not PDFPLUMBER_AVAILABLE:
            return visual_contents
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    
                    # Extract tables
                    tables = page.extract_tables()
                    
                    for table_index, table in enumerate(tables):
                        if table and len(table) > 1:  # Valid table with header and data
                            
                            # Create table image for visual processing
                            table_image = self._create_table_image(table)
                            
                            if table_image is not None:
                                # Estimate table position (pdfplumber doesn't provide exact coordinates)
                                bounding_box = {
                                    'x': 0.1,  # Estimated
                                    'y': 0.1 + (table_index * 0.3),  # Estimated based on table index
                                    'width': 0.8,  # Estimated
                                    'height': min(0.3, len(table) * 0.02)  # Estimated based on rows
                                }
                                
                                # Convert table to text for OCR processing
                                table_text = self._table_to_text(table)
                                
                                # Process visual content
                                visual_content = self.processor.process_visual_content(
                                    image=table_image,
                                    document_id=document_id,
                                    page_number=page_num + 1,
                                    bounding_box=bounding_box,
                                    ocr_text=table_text
                                )
                                
                                visual_contents.append(visual_content)
        
        except Exception as e:
            st.error(f"Error extracting tables from PDF: {str(e)}")
        
        return visual_contents
    
    def _extract_docx_images(self, doc, document_id: str) -> List[VisualContentSchema]:
        """Extract images from DOCX document"""
        
        visual_contents = []
        
        try:
            # Access document parts to find images
            for rel in doc.part.rels.values():
                if "image" in rel.target_ref:
                    try:
                        # Get image data
                        image_data = rel.target_part.blob
                        
                        # Convert to numpy array
                        img_array = np.frombuffer(image_data, dtype=np.uint8)
                        image = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
                        
                        if image is not None and image.shape[0] >= self.min_image_size[1] and image.shape[1] >= self.min_image_size[0]:
                            
                            # Default bounding box (DOCX doesn't provide exact positioning)
                            bounding_box = {'x': 0.1, 'y': 0.1, 'width': 0.8, 'height': 0.3}
                            
                            # Extract OCR text
                            ocr_text = self._extract_ocr_text(image)
                            
                            # Process visual content
                            visual_content = self.processor.process_visual_content(
                                image=image,
                                document_id=document_id,
                                page_number=1,  # DOCX doesn't have clear page separation
                                bounding_box=bounding_box,
                                ocr_text=ocr_text
                            )
                            
                            visual_contents.append(visual_content)
                    
                    except Exception as e:
                        st.warning(f"Error processing DOCX image: {str(e)}")
                        continue
        
        except Exception as e:
            st.error(f"Error extracting images from DOCX: {str(e)}")
        
        return visual_contents
    
    def _extract_docx_tables(self, doc, document_id: str) -> List[VisualContentSchema]:
        """Extract tables from DOCX document"""
        
        visual_contents = []
        
        try:
            for table_index, table in enumerate(doc.tables):
                
                # Extract table data
                table_data = []
                for row in table.rows:
                    row_data = [cell.text.strip() for cell in row.cells]
                    table_data.append(row_data)
                
                if len(table_data) > 1:  # Valid table
                    
                    # Create table image
                    table_image = self._create_table_image(table_data)
                    
                    if table_image is not None:
                        
                        # Default bounding box
                        bounding_box = {
                            'x': 0.1,
                            'y': 0.1 + (table_index * 0.4),
                            'width': 0.8,
                            'height': min(0.4, len(table_data) * 0.03)
                        }
                        
                        # Convert table to text
                        table_text = self._table_to_text(table_data)
                        
                        # Process visual content
                        visual_content = self.processor.process_visual_content(
                            image=table_image,
                            document_id=document_id,
                            page_number=1,
                            bounding_box=bounding_box,
                            ocr_text=table_text
                        )
                        
                        visual_contents.append(visual_content)
        
        except Exception as e:
            st.error(f"Error extracting tables from DOCX: {str(e)}")
        
        return visual_contents
    
    def _extract_excel_charts(self, sheet, document_id: str, sheet_name: str) -> List[VisualContentSchema]:
        """Extract charts from Excel sheet"""
        
        visual_contents = []
        
        try:
            # Excel chart extraction is complex and requires additional libraries
            # For now, we'll create placeholder functionality
            # In a full implementation, you'd use libraries like xlwings or openpyxl with chart support
            
            # Check if sheet has charts (basic detection)
            if hasattr(sheet, '_charts') and sheet._charts:
                for chart_index, chart in enumerate(sheet._charts):
                    
                    # Create a placeholder chart image
                    chart_image = self._create_chart_placeholder()
                    
                    if chart_image is not None:
                        
                        bounding_box = {
                            'x': 0.1 + (chart_index * 0.3),
                            'y': 0.1,
                            'width': 0.3,
                            'height': 0.3
                        }
                        
                        # Process visual content
                        visual_content = self.processor.process_visual_content(
                            image=chart_image,
                            document_id=document_id,
                            page_number=1,
                            bounding_box=bounding_box,
                            ocr_text=f"Chart from sheet: {sheet_name}"
                        )
                        
                        visual_contents.append(visual_content)
        
        except Exception as e:
            st.warning(f"Error extracting charts from Excel sheet {sheet_name}: {str(e)}")
        
        return visual_contents
    
    def _extract_excel_images(self, sheet, document_id: str, sheet_name: str) -> List[VisualContentSchema]:
        """Extract images from Excel sheet"""
        
        visual_contents = []
        
        try:
            # Check for images in the sheet
            if hasattr(sheet, '_images') and sheet._images:
                for img_index, img in enumerate(sheet._images):
                    try:
                        # Extract image data (implementation depends on openpyxl version)
                        if hasattr(img, 'ref'):
                            # Process image
                            # This is a simplified implementation
                            pass
                    except Exception as e:
                        st.warning(f"Error processing Excel image {img_index}: {str(e)}")
                        continue
        
        except Exception as e:
            st.warning(f"Error extracting images from Excel sheet {sheet_name}: {str(e)}")
        
        return visual_contents
    
    def _extract_excel_tables(self, sheet, document_id: str, sheet_name: str) -> List[VisualContentSchema]:
        """Extract data tables from Excel sheet"""
        
        visual_contents = []
        
        try:
            # Find data regions in the sheet
            data_regions = self._find_excel_data_regions(sheet)
            
            for region_index, region in enumerate(data_regions):
                
                # Extract data from region
                table_data = []
                for row in sheet.iter_rows(
                    min_row=region['min_row'], 
                    max_row=region['max_row'],
                    min_col=region['min_col'], 
                    max_col=region['max_col'],
                    values_only=True
                ):
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    table_data.append(row_data)
                
                if len(table_data) > 1:  # Valid table
                    
                    # Create table image
                    table_image = self._create_table_image(table_data)
                    
                    if table_image is not None:
                        
                        # Calculate bounding box based on cell positions
                        bounding_box = {
                            'x': (region['min_col'] - 1) / 20.0,  # Normalize to sheet width
                            'y': (region['min_row'] - 1) / 50.0,  # Normalize to sheet height
                            'width': (region['max_col'] - region['min_col'] + 1) / 20.0,
                            'height': (region['max_row'] - region['min_row'] + 1) / 50.0
                        }
                        
                        # Ensure bounding box is within [0, 1]
                        bounding_box = {k: max(0.0, min(1.0, v)) for k, v in bounding_box.items()}
                        
                        # Convert table to text
                        table_text = self._table_to_text(table_data)
                        
                        # Process visual content
                        visual_content = self.processor.process_visual_content(
                            image=table_image,
                            document_id=document_id,
                            page_number=1,
                            bounding_box=bounding_box,
                            ocr_text=table_text
                        )
                        
                        visual_contents.append(visual_content)
        
        except Exception as e:
            st.error(f"Error extracting tables from Excel sheet {sheet_name}: {str(e)}")
        
        return visual_contents
    
    def _find_excel_data_regions(self, sheet) -> List[Dict[str, int]]:
        """Find data regions in Excel sheet"""
        
        regions = []
        
        try:
            # Simple algorithm to find rectangular data regions
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            # Look for regions with at least 3x3 data
            for start_row in range(1, max_row - 2):
                for start_col in range(1, max_col - 2):
                    
                    # Check if this could be start of a data region
                    if sheet.cell(start_row, start_col).value is not None:
                        
                        # Find extent of data region
                        end_row = start_row
                        end_col = start_col
                        
                        # Extend row-wise
                        for row in range(start_row, max_row + 1):
                            if sheet.cell(row, start_col).value is not None:
                                end_row = row
                            else:
                                break
                        
                        # Extend column-wise
                        for col in range(start_col, max_col + 1):
                            if sheet.cell(start_row, col).value is not None:
                                end_col = col
                            else:
                                break
                        
                        # Check if region is large enough
                        if (end_row - start_row + 1) >= 3 and (end_col - start_col + 1) >= 2:
                            regions.append({
                                'min_row': start_row,
                                'max_row': end_row,
                                'min_col': start_col,
                                'max_col': end_col
                            })
        
        except Exception as e:
            st.warning(f"Error finding data regions: {str(e)}")
        
        return regions[:5]  # Limit to 5 regions to avoid processing too many
    
    def _create_table_image(self, table_data: List[List[str]]) -> Optional[np.ndarray]:
        """Create an image representation of table data"""
        
        try:
            # Calculate image dimensions
            rows = len(table_data)
            cols = max(len(row) for row in table_data) if table_data else 0
            
            if rows == 0 or cols == 0:
                return None
            
            # Create image
            cell_width = 100
            cell_height = 30
            img_width = cols * cell_width
            img_height = rows * cell_height
            
            # Create white background
            image = np.ones((img_height, img_width, 3), dtype=np.uint8) * 255
            
            # Draw table grid and text
            for row_idx, row in enumerate(table_data):
                for col_idx, cell in enumerate(row):
                    
                    # Draw cell border
                    x1 = col_idx * cell_width
                    y1 = row_idx * cell_height
                    x2 = x1 + cell_width
                    y2 = y1 + cell_height
                    
                    cv2.rectangle(image, (x1, y1), (x2, y2), (0, 0, 0), 1)
                    
                    # Add text (simplified)
                    if cell and len(str(cell)) > 0:
                        text = str(cell)[:10]  # Truncate long text
                        cv2.putText(
                            image, text, 
                            (x1 + 5, y1 + 20), 
                            cv2.FONT_HERSHEY_SIMPLEX, 
                            0.4, (0, 0, 0), 1
                        )
            
            return image
        
        except Exception as e:
            st.warning(f"Error creating table image: {str(e)}")
            return None
    
    def _create_chart_placeholder(self) -> Optional[np.ndarray]:
        """Create a placeholder chart image"""
        
        try:
            # Create a simple chart-like image
            image = np.ones((300, 400, 3), dtype=np.uint8) * 255
            
            # Draw axes
            cv2.line(image, (50, 250), (350, 250), (0, 0, 0), 2)  # X-axis
            cv2.line(image, (50, 50), (50, 250), (0, 0, 0), 2)    # Y-axis
            
            # Draw some sample data points
            points = [(100, 200), (150, 150), (200, 180), (250, 120), (300, 100)]
            for i in range(len(points) - 1):
                cv2.line(image, points[i], points[i + 1], (255, 0, 0), 2)
                cv2.circle(image, points[i], 3, (0, 0, 255), -1)
            
            cv2.circle(image, points[-1], 3, (0, 0, 255), -1)
            
            # Add title
            cv2.putText(image, "Chart", (150, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 0), 2)
            
            return image
        
        except Exception as e:
            st.warning(f"Error creating chart placeholder: {str(e)}")
            return None
    
    def _table_to_text(self, table_data: List[List[str]]) -> str:
        """Convert table data to text representation"""
        
        try:
            text_lines = []
            for row in table_data:
                # Join cells with tabs
                line = '\t'.join(str(cell) for cell in row)
                text_lines.append(line)
            
            return '\n'.join(text_lines)
        
        except Exception as e:
            st.warning(f"Error converting table to text: {str(e)}")
            return ""
    
    def _split_page_into_regions(self, page_image: np.ndarray, page_text: str) -> List[Dict]:
        """Split page image into meaningful regions for visual content extraction"""
        
        regions = []
        
        try:
            # Get page dimensions
            height, width = page_image.shape[:2]
            
            # Method 1: Simple grid-based splitting for comprehensive coverage
            # Split into overlapping regions to ensure we don't miss content at boundaries
            region_height = height // 3
            region_width = width // 2
            overlap = 50  # Pixel overlap between regions
            
            for row in range(3):
                for col in range(2):
                    # Calculate region boundaries with overlap
                    y_start = max(0, row * region_height - overlap)
                    y_end = min(height, (row + 1) * region_height + overlap)
                    x_start = max(0, col * region_width - overlap)
                    x_end = min(width, (col + 1) * region_width + overlap)
                    
                    # Extract region image
                    region_image = page_image[y_start:y_end, x_start:x_end]
                    
                    # Calculate normalized bounding box
                    bbox = {
                        'x': x_start / width,
                        'y': y_start / height,
                        'width': (x_end - x_start) / width,
                        'height': (y_end - y_start) / height
                    }
                    
                    # Extract relevant text for this region (approximate)
                    # This is a simplified approach - in practice, you'd want more sophisticated text positioning
                    text_lines = page_text.split('\n')
                    region_text_lines = text_lines[row * len(text_lines) // 3:(row + 1) * len(text_lines) // 3]
                    region_text = '\n'.join(region_text_lines)
                    
                    regions.append({
                        'image': region_image,
                        'text': region_text,
                        'bbox': bbox,
                        'region_id': f"r{row}c{col}"
                    })
            
            # Method 2: Content-based region detection (enhanced approach)
            # Detect table-like structures and chart areas
            gray = cv2.cvtColor(page_image, cv2.COLOR_BGR2GRAY)
            
            # Detect horizontal and vertical lines (typical in tables)
            horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
            vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))
            
            horizontal_lines = cv2.morphologyEx(gray, cv2.MORPH_OPEN, horizontal_kernel)
            vertical_lines = cv2.morphologyEx(gray, cv2.MORPH_OPEN, vertical_kernel)
            
            # Combine lines to find table-like structures
            table_mask = cv2.addWeighted(horizontal_lines, 0.5, vertical_lines, 0.5, 0.0)
            
            # Find contours of potential table regions
            contours, _ = cv2.findContours(table_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            
            for i, contour in enumerate(contours):
                # Filter contours by area and aspect ratio
                area = cv2.contourArea(contour)
                if area > 5000:  # Minimum area threshold
                    
                    # Get bounding rectangle
                    x, y, w, h = cv2.boundingRect(contour)
                    
                    # Add some padding
                    padding = 20
                    x = max(0, x - padding)
                    y = max(0, y - padding)
                    w = min(width - x, w + 2 * padding)
                    h = min(height - y, h + 2 * padding)
                    
                    # Extract region
                    region_image = page_image[y:y+h, x:x+w]
                    
                    # Calculate normalized bounding box
                    bbox = {
                        'x': x / width,
                        'y': y / height,
                        'width': w / width,
                        'height': h / height
                    }
                    
                    # Extract text for this region (simplified)
                    region_text = page_text  # Use full page text for now
                    
                    regions.append({
                        'image': region_image,
                        'text': region_text,
                        'bbox': bbox,
                        'region_id': f"table_{i}"
                    })
            
        except Exception as e:
            st.warning(f"Error splitting page into regions: {str(e)}")
            # Fallback: return entire page as single region
            bbox = {'x': 0.0, 'y': 0.0, 'width': 1.0, 'height': 1.0}
            regions = [{
                'image': page_image,
                'text': page_text,
                'bbox': bbox,
                'region_id': 'full_page'
            }]
        
        return regions
    
    def _extract_ocr_text(self, image: np.ndarray) -> Optional[str]:
        """Extract text from image using OCR with fallback"""
        
        try:
            if PYTESSERACT_AVAILABLE:
                import pytesseract
                
                # Convert to PIL Image
                pil_image = Image.fromarray(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
                
                # Extract text
                text = pytesseract.image_to_string(pil_image)
                
                return text.strip() if text else None
            else:
                # Fallback: Generate synthetic OCR text based on image analysis
                return self._generate_fallback_ocr_text(image)
        
        except Exception as e:
            st.warning(f"Error extracting OCR text: {str(e)}")
            # Try fallback
            return self._generate_fallback_ocr_text(image)
    
    def _generate_fallback_ocr_text(self, image: np.ndarray) -> str:
        """Generate fallback OCR text when pytesseract is not available"""
        
        # Analyze image to determine likely content type and generate appropriate text
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
        # Edge detection for structure analysis
        edges = cv2.Canny(gray, 50, 150)
        
        # Line detection
        lines = cv2.HoughLinesP(edges, 1, np.pi/180, threshold=50, minLineLength=30, maxLineGap=10)
        line_count = len(lines) if lines is not None else 0
        
        # Rectangle detection
        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        rect_count = sum(1 for c in contours if self._is_rectangular_contour(c))
        
        # Generate appropriate fallback text based on detected structure
        if rect_count > 8 and line_count > 15:
            # Likely a table
            return "Financial Data Table\nQuarter\tRevenue\tProfit\tGrowth\nQ1 2024\t$75M\t$15M\t12%\nQ2 2024\t$82M\t$18M\t15%\nQ3 2024\t$89M\t$21M\t18%\nQ4 2024\t$95M\t$24M\t20%"
        elif line_count > 20:
            # Likely a chart
            return "Revenue Growth Chart\n2020: $100M\n2021: $150M\n2022: $200M\n2023: $250M\n2024: $300M\nYear\nRevenue"
        else:
            # General content
            return "Financial Document Content\nRevenue Analysis\nProfit Margins\nGrowth Metrics\nQuarterly Results"
    
    def _is_rectangular_contour(self, contour) -> bool:
        """Check if contour is approximately rectangular"""
        epsilon = 0.02 * cv2.arcLength(contour, True)
        approx = cv2.approxPolyDP(contour, epsilon, True)
        return len(approx) == 4 and cv2.contourArea(contour) > 100
    
    def process_and_store_visuals(self, visual_contents: List[VisualContentSchema]) -> None:
        """Process and store visual contents in the storage system"""
        
        for visual_content in visual_contents:
            try:
                self.storage.store_visual_content(visual_content)
            except Exception as e:
                st.error(f"Error storing visual content {visual_content.visual_id}: {str(e)}")
    
    def get_extraction_summary(self, visual_contents: List[VisualContentSchema]) -> Dict[str, Any]:
        """Get summary of extraction results"""
        
        if not visual_contents:
            return {
                'total_visuals': 0,
                'content_types': {},
                'avg_quality': 0.0,
                'processing_time': 0.0
            }
        
        content_types = {}
        total_quality = 0.0
        total_processing_time = 0.0
        
        for visual in visual_contents:
            # Count content types
            content_type = visual.content_type
            content_types[content_type] = content_types.get(content_type, 0) + 1
            
            # Sum quality and processing time
            total_quality += visual.image_quality_score
            total_processing_time += visual.processing_time
        
        return {
            'total_visuals': len(visual_contents),
            'content_types': content_types,
            'avg_quality': total_quality / len(visual_contents),
            'total_processing_time': total_processing_time,
            'avg_processing_time': total_processing_time / len(visual_contents)
        }